from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from datetime import datetime, timedelta
from django.db.models import F, Sum, Count
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, Extract
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, DateFromToRangeFilter, NumberFilter, CharFilter, ChoiceFilter
from django.db.models import Q
from django.db import transaction, models
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.conf import settings
import json
import csv
import io
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)

from .models import Proforma, ProformaItem, ProformaHistorial, ConfiguracionProforma
from .serializers import (
    ProformaSerializer, ProformaItemSerializer, 
    ProformaHistorialSerializer, ConfiguracionProformaSerializer,
    BusquedaProductosSerializer
)
from .pagination import StandardResultsSetPagination, LargeResultsSetPagination

from pandora.models import Clientes, EmpresaClc
from products.models import ProductoOfertado, ProductoDisponible

# Definición de filtro personalizado para Proformas
class ProformaFilter(FilterSet):
    """Filtro personalizado para el modelo Proforma con funcionalidades avanzadas"""
    # Filtros de búsqueda por texto
    search = CharFilter(method='filter_search', label='Búsqueda global')
    nombre_contains = CharFilter(field_name='nombre', lookup_expr='icontains')
    numero_contains = CharFilter(field_name='numero', lookup_expr='icontains')
    
    # Filtros de fecha
    fecha_emision = DateFromToRangeFilter()
    fecha_vencimiento = DateFromToRangeFilter()
    created_at = DateFromToRangeFilter()
    
    # Filtros de relaciones
    cliente_nombre = CharFilter(field_name='cliente__nombre', lookup_expr='icontains')
    empresa_nombre = CharFilter(field_name='empresa__nombre', lookup_expr='icontains')
    
    # Filtros de rango para montos
    subtotal_min = NumberFilter(field_name='subtotal', lookup_expr='gte')
    subtotal_max = NumberFilter(field_name='subtotal', lookup_expr='lte')
    total_min = NumberFilter(field_name='total', lookup_expr='gte')
    total_max = NumberFilter(field_name='total', lookup_expr='lte')
    
    # Filtro de estado específico
    estado = ChoiceFilter(choices=Proforma.ESTADO_CHOICES)
    
    # Filtro para múltiples estados
    estados = CharFilter(method='filter_estados', label='Múltiples estados separados por coma')
    
    class Meta:
        model = Proforma
        fields = [
            'search', 'nombre_contains', 'numero_contains',
            'fecha_emision', 'fecha_vencimiento', 'created_at',
            'cliente', 'cliente_nombre', 'empresa', 'empresa_nombre',
            'subtotal_min', 'subtotal_max', 'total_min', 'total_max',
            'estado', 'estados', 'created_by'
        ]
    
    def filter_search(self, queryset, name, value):
        """Búsqueda global en múltiples campos de la proforma"""
        if not value:
            return queryset
            
        # Buscar en múltiples campos relevantes
        query = Q(numero__icontains=value) | \
                Q(nombre__icontains=value) | \
                Q(notas__icontains=value) | \
                Q(cliente__nombre__icontains=value) | \
                Q(cliente__ruc__icontains=value) | \
                Q(atencion_a__icontains=value)
                
        return queryset.filter(query).distinct()
    
    def filter_estados(self, queryset, name, value):
        """Permite filtrar por múltiples estados separados por coma"""
        if not value:
            return queryset
            
        # Dividir la cadena en estados individuales y filtrar
        estados = [estado.strip() for estado in value.split(',')]
        return queryset.filter(estado__in=estados)


# Filtro personalizado para ítems de proforma
class ProformaItemFilter(FilterSet):
    """Filtro para ítems de proforma"""
    descripcion_contains = CharFilter(field_name='descripcion', lookup_expr='icontains')
    codigo_contains = CharFilter(field_name='codigo', lookup_expr='icontains')
    precio_min = NumberFilter(field_name='precio_unitario', lookup_expr='gte')
    precio_max = NumberFilter(field_name='precio_unitario', lookup_expr='lte')
    
    class Meta:
        model = ProformaItem
        fields = [
            'proforma', 'tipo_item', 'descripcion_contains', 'codigo_contains',
            'precio_min', 'precio_max', 'producto_ofertado', 'producto_disponible'
        ]


class ProformaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar proformas
    """
    queryset = Proforma.objects.select_related(
        'cliente', 'empresa', 'tipo_contratacion', 'created_by', 'updated_by'
    ).prefetch_related('items', 'historial').all()
    serializer_class = ProformaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProformaFilter
    search_fields = ['numero', 'nombre', 'cliente__nombre', 'notas', 'atencion_a']
    ordering_fields = ['numero', 'fecha_emision', 'fecha_vencimiento', 'cliente__nombre', 'total', 'estado', 'created_at']
    ordering = ['-fecha_emision']
    pagination_class = StandardResultsSetPagination
    
    @transaction.atomic
    def perform_create(self, serializer):
        """Asignar el usuario actual como creador dentro de una transacción atómica"""
        try:
            serializer.save(created_by=self.request.user, updated_by=self.request.user)
            logger.info(f"Proforma creada por usuario {self.request.user.username}")
        except Exception as e:
            logger.exception(f"Error al crear proforma: {str(e)}")
            raise
    
    @transaction.atomic
    def perform_update(self, serializer):
        """Asignar el usuario actual como actualizador dentro de una transacción atómica"""
        try:
            serializer.save(updated_by=self.request.user)
            logger.info(f"Proforma {serializer.instance.id} actualizada por usuario {self.request.user.username}")
        except Exception as e:
            logger.exception(f"Error al actualizar proforma {serializer.instance.id}: {str(e)}")
            raise
    
    @action(detail=True, methods=['get'])
    def items(self, request, pk=None):
        """Obtener los ítems de una proforma específica"""
        proforma = self.get_object()
        items = proforma.items.all()
        serializer = ProformaItemSerializer(items, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Obtener el historial de una proforma específica"""
        proforma = self.get_object()
        historial = proforma.historial.all()
        serializer = ProformaHistorialSerializer(historial, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def cambiar_estado(self, request, pk=None):
        """Cambiar el estado de una proforma"""
        try:
            proforma = self.get_object()
            nuevo_estado = request.data.get('estado', None)
            
            if not nuevo_estado:
                return Response(
                    {"error": "Debe proporcionar un nuevo estado"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            estados_validos = dict(Proforma.ESTADO_CHOICES).keys()
            if nuevo_estado not in estados_validos:
                return Response(
                    {"error": f"Estado no válido. Opciones: {', '.join(estados_validos)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Almacenar estado anterior
            estado_anterior = proforma.estado
            
            # Cambiar estado
            proforma.estado = nuevo_estado
            proforma.save(update_fields=['estado'])
            
            # El registro en historial ahora se hace automáticamente via signals
            
            logger.info(
                f"Proforma #{proforma.numero} cambio de estado: {estado_anterior} -> {nuevo_estado} "  
                f"por usuario: {request.user.username}"
            )
            
            # Devolver proforma actualizada
            serializer = self.get_serializer(proforma)
            return Response(serializer.data)
            
        except Exception as e:
            logger.exception(f"Error al cambiar estado de proforma {pk}: {str(e)}")
            return Response(
                {"error": f"No se pudo cambiar el estado: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def enviar(self, request, pk=None):
        """Marcar una proforma como enviada al cliente usando el método de transición"""
        try:
            proforma = self.get_object()
            
            # Usar el método de transición definido en el modelo
            proforma.enviar(
                usuario=request.user,
                notas=request.data.get('notas', '')
            )
            
            # Devolver proforma actualizada
            serializer = self.get_serializer(proforma)
            return Response(serializer.data)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.exception(f"Error al enviar proforma {pk}: {str(e)}")
            return Response(
                {"error": f"Error al enviar proforma: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None):
        """Marcar una proforma como aprobada por el cliente usando el método de transición"""
        try:
            proforma = self.get_object()
            
            # Usar el método de transición definido en el modelo
            proforma.aprobar(
                usuario=request.user,
                notas=request.data.get('notas', '')
            )
            
            # Devolver proforma actualizada
            serializer = self.get_serializer(proforma)
            return Response(serializer.data)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.exception(f"Error al aprobar proforma {pk}: {str(e)}")
            return Response(
                {"error": f"Error al aprobar proforma: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def rechazar(self, request, pk=None):
        """Marcar una proforma como rechazada por el cliente usando el método de transición"""
        try:
            proforma = self.get_object()
            
            # Usar el método de transición definido en el modelo
            proforma.rechazar(
                usuario=request.user,
                notas=request.data.get('notas', '')
            )
            
            # Devolver proforma actualizada
            serializer = self.get_serializer(proforma)
            return Response(serializer.data)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.exception(f"Error al rechazar proforma {pk}: {str(e)}")
            return Response(
                {"error": f"Error al rechazar proforma: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    @action(detail=True, methods=['post'])
    def convertir(self, request, pk=None):
        """Convertir proforma a orden usando el método de transición"""
        try:
            proforma = self.get_object()
            
            # Usar el método de transición definido en el modelo
            proforma.convertir(
                usuario=request.user,
                notas=request.data.get('notas', '')
            )
            
            # Devolver proforma actualizada
            serializer = self.get_serializer(proforma)
            return Response(serializer.data)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.exception(f"Error al convertir proforma {pk}: {str(e)}")
            return Response(
                {"error": f"Error al convertir proforma: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def volver_a_borrador(self, request, pk=None):
        """Volver a marcar proforma como borrador usando el método de transición"""
        try:
            proforma = self.get_object()
            
            # Usar el método de transición definido en el modelo
            proforma.volver_a_borrador(
                usuario=request.user,
                notas=request.data.get('notas', '')
            )
            
            # Devolver proforma actualizada
            serializer = self.get_serializer(proforma)
            return Response(serializer.data)
            
        except ValueError as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.exception(f"Error al revertir proforma {pk} a borrador: {str(e)}")
            return Response(
                {"error": f"Error al revertir proforma a borrador: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @transaction.atomic
    @action(detail=True, methods=['post'])
    def duplicar(self, request, pk=None):
        """Crear una copia de una proforma existente"""
        try:
            proforma_original = self.get_object()
            
            # Crear nueva proforma
            nueva_proforma = Proforma.objects.create(
                numero=Proforma().generar_numero(),
                fecha_emision=timezone.now().date(),
                fecha_vencimiento=timezone.now().date() + timezone.timedelta(days=15),
                cliente=proforma_original.cliente,
                empresa=proforma_original.empresa,
                tipo_contratacion=proforma_original.tipo_contratacion,
                atencion_a=proforma_original.atencion_a,
                condiciones_pago=proforma_original.condiciones_pago,
                tiempo_entrega=proforma_original.tiempo_entrega,
                porcentaje_impuesto=proforma_original.porcentaje_impuesto,
                notas=proforma_original.notas,
                estado='borrador',
                created_by=request.user,
                updated_by=request.user
            )
            
            # Duplicar ítems con bulk_create
            items_originales = proforma_original.items.all()
            nuevos_items = []
            
            for item_original in items_originales:
                nuevos_items.append(
                    ProformaItem(
                        proforma=nueva_proforma,
                        tipo_item=item_original.tipo_item,
                        producto_ofertado=item_original.producto_ofertado,
                        producto_disponible=item_original.producto_disponible,
                        codigo=item_original.codigo,
                        descripcion=item_original.descripcion,
                        unidad=item_original.unidad,
                        cantidad=item_original.cantidad,
                        precio_unitario=item_original.precio_unitario,
                        porcentaje_descuento=item_original.porcentaje_descuento,
                        total=item_original.total,
                        orden=item_original.orden
                    )
                )
            
            # Usar bulk_create para optimizar la inserción masiva de ítems
            if nuevos_items:
                ProformaItem.objects.bulk_create(nuevos_items)
                logger.info(f"Se duplicaron {len(nuevos_items)} ítems para la proforma #{nueva_proforma.numero}")
            
            # Calcular totales
            nueva_proforma.calcular_montos()
            nueva_proforma.save(update_fields=['subtotal', 'impuesto', 'total'])
            
            # Registrar en historial (ahora automático con signals)
            
            # Devolver nueva proforma
            serializer = self.get_serializer(nueva_proforma)
            return Response(serializer.data)
            
        except Exception as e:
            logger.exception(f"Error al duplicar proforma {pk}: {str(e)}")
            return Response(
                {"error": f"No se pudo duplicar la proforma: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def reporte_estadisticas(self, request):
        """
        Generar un reporte de estadísticas en Excel
        
        Parámetros de filtrado disponibles:
        - estado: Estado de las proformas ('borrador', 'enviada', etc.) o múltiples estados separados por coma
        - cliente_id: ID del cliente
        - fecha_inicio: Fecha de inicio (formato: YYYY-MM-DD)
        - fecha_fin: Fecha fin (formato: YYYY-MM-DD)
        - vendedor_id: ID del usuario que creó las proformas
        - incluir_ventas: Si se incluyen estadísticas de ventas (conversiones)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference, PieChart, LineChart
            from django.db.models.functions import TruncMonth, TruncWeek, TruncDay
            from django.db.models import Sum, Count, F, ExpressionWrapper, FloatField, Avg, Case, When, Value, IntegerField
            
            # Obtener parámetros de filtrado
            estado = request.query_params.get('estado', None)
            cliente_id = request.query_params.get('cliente_id', None)
            fecha_inicio = request.query_params.get('fecha_inicio', None)
            fecha_fin = request.query_params.get('fecha_fin', None)
            vendedor_id = request.query_params.get('vendedor_id', None)
            incluir_ventas = request.query_params.get('incluir_ventas', 'false').lower() == 'true'
            formato_tiempo = request.query_params.get('formato_tiempo', 'month') # 'month', 'week', 'day'
            
            # Base de datos de proformas
            queryset = Proforma.objects.select_related('cliente', 'empresa', 'created_by').all()
            
            # Aplicar filtros si existen
            if estado:
                if ',' in estado:
                    estados = [e.strip() for e in estado.split(',')]
                    queryset = queryset.filter(estado__in=estados)
                else:
                    queryset = queryset.filter(estado=estado)
            
            if cliente_id:
                queryset = queryset.filter(cliente_id=cliente_id)
                
            if fecha_inicio:
                queryset = queryset.filter(fecha_emision__gte=fecha_inicio)
                
            if fecha_fin:
                queryset = queryset.filter(fecha_emision__lte=fecha_fin)
            
            if vendedor_id:
                queryset = queryset.filter(created_by_id=vendedor_id)
            
            # Crear workbook y hojas
            wb = openpyxl.Workbook()
            ws_resumen = wb.active
            ws_resumen.title = "Resumen"
            
            # Estilos
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            subheader_font = Font(bold=True, size=11)
            
            # Título y fecha
            ws_resumen.merge_cells('A1:G1')
            title_cell = ws_resumen.cell(row=1, column=1, value="REPORTE DE ESTADÍSTICAS DE PROFORMAS")
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center")
            
            fecha_actual = timezone.now().strftime("%d/%m/%Y %H:%M")
            ws_resumen.cell(row=2, column=1, value=f"Generado: {fecha_actual}")
            
            # Información del usuario que genera el reporte
            if request.user and hasattr(request.user, 'username'):
                ws_resumen.cell(row=2, column=4, value=f"Generado por: {request.user.get_full_name() or request.user.username}")
            
            # Filtros aplicados
            ws_resumen.cell(row=4, column=1, value="Filtros aplicados:").font = Font(bold=True)
            filters = []
            
            if estado:
                if ',' in estado:
                    estados = ", ".join([dict(Proforma.ESTADO_CHOICES).get(e.strip(), e.strip()) for e in estado.split(',')])
                    filters.append(f"Estados: {estados}")
                else:
                    estado_display = dict(Proforma.ESTADO_CHOICES).get(estado, estado)
                    filters.append(f"Estado: {estado_display}")
            
            if cliente_id:
                try:
                    cliente = Clientes.objects.get(pk=cliente_id)
                    filters.append(f"Cliente: {cliente.nombre}")
                except Exception as e:
                    filters.append(f"Cliente ID: {cliente_id}")
            
            if fecha_inicio:
                filters.append(f"Desde: {fecha_inicio}")
                
            if fecha_fin:
                filters.append(f"Hasta: {fecha_fin}")
            
            if vendedor_id:
                try:
                    from django.contrib.auth import get_user_model
                    User = get_user_model()
                    vendedor = User.objects.get(pk=vendedor_id)
                    filters.append(f"Vendedor: {vendedor.get_full_name() or vendedor.username}")
                except Exception as e:
                    filters.append(f"Vendedor ID: {vendedor_id}")
                
            if not filters:
                filters.append("Ninguno (mostrando todas las proformas)")
                
            for i, filter_text in enumerate(filters, 5):
                ws_resumen.cell(row=i, column=1, value=filter_text)
            
            # Resumen general
            ws_resumen.cell(row=7, column=1, value="RESUMEN GENERAL").font = subheader_font
            
            # Total de proformas y montos
            total_proformas = queryset.count()
            monto_total = sum(queryset.values_list('total', flat=True)) or 0
            
            # Datos estadísticos generales
            monto_promedio = monto_total / total_proformas if total_proformas > 0 else 0
            aprobadas = queryset.filter(estado='aprobada').count()
            convertidas = queryset.filter(estado='convertida').count()
            tasa_aprobacion = (aprobadas / total_proformas * 100) if total_proformas > 0 else 0
            tasa_conversion = (convertidas / total_proformas * 100) if total_proformas > 0 else 0
            
            # Mostrar datos generales en formato de tabla
            resumen_data = [
                ("Total de proformas:", total_proformas, "", "Monto total:", f"${monto_total:,.2f}"),
                ("Proformas aprobadas:", aprobadas, "", "Monto promedio:", f"${monto_promedio:,.2f}"),
                ("Proformas convertidas:", convertidas, "", "Tasa de aprobación:", f"{tasa_aprobacion:.2f}%"),
                ("", "", "", "Tasa de conversión:", f"{tasa_conversion:.2f}%"),
            ]
            
            for i, row_data in enumerate(resumen_data, 9):
                for j, value in enumerate(row_data, 1):
                    # Aplicar formato en negrita para las etiquetas
                    if j in [1, 4]:  # Columnas de etiquetas
                        ws_resumen.cell(row=i, column=j, value=value).font = Font(bold=True)
                    else:
                        ws_resumen.cell(row=i, column=j, value=value)
            
            # Total de proformas por estado
            ws_resumen.cell(row=15, column=1, value="PROFORMAS POR ESTADO").font = subheader_font
            
            headers = ["Estado", "Cantidad", "% del Total", "Monto Total", "% del Monto Total", "Precio Promedio"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_resumen.cell(row=16, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                ws_resumen.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 2)
                
            # Obtener estadísticas por estado
            estadisticas_estado = []
            
            for estado, label in Proforma.ESTADO_CHOICES:
                state_queryset = queryset.filter(estado=estado)
                count = state_queryset.count()
                total = sum(state_queryset.values_list('total', flat=True)) or 0
                
                if count > 0 or total > 0:
                    estadisticas_estado.append({
                        'estado': estado,
                        'label': label,
                        'count': count,
                        'total': total,
                        'promedio': total / count if count > 0 else 0,
                        'porcentaje_count': (count / total_proformas * 100) if total_proformas > 0 else 0,
                        'porcentaje_total': (total / monto_total * 100) if monto_total > 0 else 0
                    })
            
            # Ordenar por cantidad, de mayor a menor
            estadisticas_estado.sort(key=lambda x: x['count'], reverse=True)
            
            # Añadir datos a la tabla
            row = 17
            for estado_stat in estadisticas_estado:
                ws_resumen.cell(row=row, column=1, value=estado_stat['label'])
                ws_resumen.cell(row=row, column=2, value=estado_stat['count'])
                
                # Porcentaje del total de proformas
                cell = ws_resumen.cell(row=row, column=3, value=estado_stat['porcentaje_count'] / 100)
                cell.number_format = '0.00%'
                
                # Monto total
                cell = ws_resumen.cell(row=row, column=4, value=estado_stat['total'])
                cell.number_format = '#,##0.00'
                
                # Porcentaje del monto total
                cell = ws_resumen.cell(row=row, column=5, value=estado_stat['porcentaje_total'] / 100)
                cell.number_format = '0.00%'
                
                # Precio promedio
                cell = ws_resumen.cell(row=row, column=6, value=estado_stat['promedio'])
                cell.number_format = '#,##0.00'
                
                row += 1
            
            # Crear gráfico de barras para proformas por estado
            chart1 = BarChart()
            chart1.title = "Proformas por Estado"
            chart1.style = 10
            chart1.x_axis.title = "Estado"
            chart1.y_axis.title = "Cantidad"
            
            data = Reference(ws_resumen, min_col=2, min_row=16, max_row=row-1, max_col=2)
            cats = Reference(ws_resumen, min_col=1, min_row=17, max_row=row-1)
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.shape = 4
            chart1.height = 10
            chart1.width = 15
            
            ws_resumen.add_chart(chart1, "H16")
            
            # Crear gráfico circular para distribución de montos
            chart2 = PieChart()
            chart2.title = "Distribución de Montos por Estado"
            chart2.style = 10
            
            data = Reference(ws_resumen, min_col=4, min_row=16, max_row=row-1, max_col=4)
            cats = Reference(ws_resumen, min_col=1, min_row=17, max_row=row-1)
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)
            chart2.height = 10
            chart2.width = 15
            
            ws_resumen.add_chart(chart2, "H32")
            
            # Proformas por cliente
            ws_clientes = wb.create_sheet(title="Por Cliente")
            
            # Título
            ws_clientes.merge_cells('A1:F1')
            title_cell = ws_clientes.cell(row=1, column=1, value="PROFORMAS POR CLIENTE")
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center")
            
            # Encabezados
            headers = ["Cliente", "RUC", "Cantidad", "Monto Total", "Proforma Promedio", "% Aprobación", "Estado Más Común"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_clientes.cell(row=3, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                ws_clientes.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 2)
            
            # Obtener estadísticas por cliente con más detalles
            clientes_stats = queryset.values(
                'cliente__id', 'cliente__nombre', 'cliente__ruc'
            ).annotate(
                count=Count('id'),
                total=Sum('total'),
                promedio=ExpressionWrapper(Sum('total') / Count('id'), output_field=FloatField()),
                aprobadas=Count('id', filter=models.Q(estado__in=['aprobada', 'convertida'])),
            ).order_by('-total')[:30]  # Limitar a los 30 clientes con mayor monto
            
            # Para cada cliente, encontrar el estado más común y calcular tasas
            for stats in clientes_stats:
                cliente_id = stats['cliente__id']
                # Encontrar el estado más común
                estados_cliente = queryset.filter(
                    cliente__id=cliente_id
                ).values('estado').annotate(
                    count=Count('id')
                ).order_by('-count')
                
                if estados_cliente:
                    estado_comun = estados_cliente[0]['estado']
                    stats['estado_comun'] = dict(Proforma.ESTADO_CHOICES).get(estado_comun, estado_comun)
                else:
                    stats['estado_comun'] = 'N/A'
                
                # Calcular tasa de aprobación
                stats['tasa_aprobacion'] = (stats['aprobadas'] / stats['count'] * 100) if stats['count'] > 0 else 0
            
            # Añadir datos a la tabla
            row = 4
            for client_stat in clientes_stats:
                ws_clientes.cell(row=row, column=1, value=client_stat['cliente__nombre'])
                ws_clientes.cell(row=row, column=2, value=client_stat['cliente__ruc'])
                ws_clientes.cell(row=row, column=3, value=client_stat['count'])
                
                # Monto total
                cell = ws_clientes.cell(row=row, column=4, value=client_stat['total'])
                cell.number_format = '#,##0.00'
                
                # Promedio
                cell = ws_clientes.cell(row=row, column=5, value=client_stat['promedio'])
                cell.number_format = '#,##0.00'
                
                # Tasa de aprobación
                cell = ws_clientes.cell(row=row, column=6, value=client_stat['tasa_aprobacion'] / 100)
                cell.number_format = '0.00%'
                
                # Estado más común
                ws_clientes.cell(row=row, column=7, value=client_stat['estado_comun'])
                
                row += 1
            
            # Crear gráfico de barras para los 10 principales clientes
            chart3 = BarChart()
            chart3.title = "Top 10 Clientes por Monto Total"
            chart3.style = 10
            chart3.x_axis.title = "Cliente"
            chart3.y_axis.title = "Monto Total"
            
            data = Reference(ws_clientes, min_col=4, min_row=3, max_row=14, max_col=4)  # Solo los primeros 10
            cats = Reference(ws_clientes, min_col=1, min_row=4, max_row=14)
            chart3.add_data(data, titles_from_data=True)
            chart3.set_categories(cats)
            chart3.shape = 4
            chart3.height = 15
            chart3.width = 20
            
            ws_clientes.add_chart(chart3, "I4")
            
            # Proformas por período de tiempo (mes, semana o día)
            ws_tiempo = wb.create_sheet(title="Por Tiempo")
            
            # Título
            ws_tiempo.merge_cells('A1:D1')
            title_cell = ws_tiempo.cell(row=1, column=1, value=f"PROFORMAS POR {formato_tiempo.upper()}")
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center")
            
            # Encabezados para tabla temporal
            ws_tiempo.cell(row=3, column=1, value=f"POR {formato_tiempo.upper()}").font = Font(bold=True)
            
            # Determinar la función de truncado según el formato seleccionado
            if formato_tiempo == 'week':
                trunc_func = TruncWeek('fecha_emision')
                period_format = "Semana %W, %Y"  # Formato de semana: "Semana 01, 2023"
                headers = ["Semana", "Año", "Cantidad", "Monto Total", "Promedio", "% Aprobadas", "Tasa Conversión"]
            elif formato_tiempo == 'day':
                trunc_func = TruncDay('fecha_emision')
                period_format = "%d/%m/%Y"  # Formato de día: "01/01/2023"
                headers = ["Día", "Mes", "Cantidad", "Monto Total", "Promedio", "% Aprobadas", "Tasa Conversión"]
            else:  # por defecto 'month'
                trunc_func = TruncMonth('fecha_emision')
                period_format = "%B %Y"  # Formato de mes: "Enero 2023"
                headers = ["Mes", "Año", "Cantidad", "Monto Total", "Promedio", "% Aprobadas", "Tasa Conversión"]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws_tiempo.cell(row=4, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                ws_tiempo.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 2)
            
            # Obtener estadísticas por período
            periodo_stats = queryset.annotate(
                periodo=trunc_func
            ).values('periodo').annotate(
                cantidad=Count('id'),
                total=Sum('total'),
                promedio=ExpressionWrapper(Sum('total') / Count('id'), output_field=FloatField()),
                aprobadas=Count('id', filter=models.Q(estado='aprobada')),
                convertidas=Count('id', filter=models.Q(estado='convertida')),
            ).order_by('-periodo')
            
            # Añadir tasas y formatear fechas para mostrar
            for stats in periodo_stats:
                periodo = stats['periodo']
                cantidad = stats['cantidad']
                
                # Calcular tasas
                stats['tasa_aprobacion'] = (stats['aprobadas'] / cantidad * 100) if cantidad > 0 else 0
                stats['tasa_conversion'] = (stats['convertidas'] / cantidad * 100) if cantidad > 0 else 0
                
                # Formatear período para mostrar
                if periodo:
                    if formato_tiempo == 'month':
                        stats['periodo_texto'] = periodo.strftime('%B')
                        stats['anio'] = periodo.year
                    elif formato_tiempo == 'week':
                        # Para semanas, mostrar el número de semana y año
                        stats['periodo_texto'] = f"Semana {periodo.strftime('%W')}"
                        stats['anio'] = periodo.year
                    else:  # day
                        stats['periodo_texto'] = periodo.strftime('%d/%m')
                        stats['anio'] = periodo.strftime('%b')
                else:
                    stats['periodo_texto'] = 'N/A'
                    stats['anio'] = 'N/A'
                    
            # Añadir datos a la tabla
            row = 5
            for periodo_stat in periodo_stats:
                ws_tiempo.cell(row=row, column=1, value=periodo_stat['periodo_texto'])
                ws_tiempo.cell(row=row, column=2, value=periodo_stat['anio'])
                ws_tiempo.cell(row=row, column=3, value=periodo_stat['cantidad'])
                
                # Monto total
                cell = ws_tiempo.cell(row=row, column=4, value=periodo_stat['total'])
                cell.number_format = '#,##0.00'
                
                # Promedio
                cell = ws_tiempo.cell(row=row, column=5, value=periodo_stat['promedio'])
                cell.number_format = '#,##0.00'
                
                # Porcentaje de aprobadas
                cell = ws_tiempo.cell(row=row, column=6, value=periodo_stat['tasa_aprobacion'] / 100)
                cell.number_format = '0.00%'
                
                # Tasa de conversión
                cell = ws_tiempo.cell(row=row, column=7, value=periodo_stat['tasa_conversion'] / 100)
                cell.number_format = '0.00%'
                
                row += 1
            
            # Crear gráfico combinado de líneas y barras para evolución temporal
            chart4 = LineChart()
            chart4.title = f"Evolución de Proformas por {formato_tiempo.capitalize()}"
            chart4.style = 10
            chart4.y_axis.title = "Monto Total"
            chart4.x_axis.title = formato_tiempo.capitalize()
            
            # Datos de monto total (línea principal)
            data = Reference(ws_tiempo, min_col=4, min_row=4, max_row=min(row-1, 24), max_col=4)  # Limitar a 24 períodos para legibilidad
            cats = Reference(ws_tiempo, min_col=1, min_row=5, max_row=min(row-1, 24))
            chart4.add_data(data, titles_from_data=True)
            chart4.set_categories(cats)
            
            # Añadir línea de cantidad como serie secundaria (eje derecho)
            data2 = Reference(ws_tiempo, min_col=3, min_row=4, max_row=min(row-1, 24), max_col=3)
            chart4.add_data(data2, titles_from_data=True)
            
            # Configurar la apariencia
            chart4.shape = 4
            chart4.height = 15
            chart4.width = 20
            
            # Añadir el gráfico a la hoja
            ws_tiempo.add_chart(chart4, "I5")
            
            # Si se solicita incluir datos de ventas, añadir una hoja específica
            if incluir_ventas:
                ws_ventas = wb.create_sheet(title="Ventas")
                
                # Título de la hoja de ventas
                ws_ventas.merge_cells('A1:E1')
                title_cell = ws_ventas.cell(row=1, column=1, value="ANÁLISIS DE CONVERSIÓN A VENTAS")
                title_cell.font = title_font
                title_cell.alignment = Alignment(horizontal="center")
                
                # Obtener proformas convertidas
                ventas_queryset = queryset.filter(estado='convertida')
                total_convertidas = ventas_queryset.count()
                monto_convertido = sum(ventas_queryset.values_list('total', flat=True)) or 0
                
                # Datos generales de conversión
                ws_ventas.cell(row=3, column=1, value="RESUMEN DE CONVERSIÓN").font = subheader_font
                
                ventas_data = [
                    ("Total proformas:", total_proformas, "", "Tasa de conversión:", f"{tasa_conversion:.2f}%"),
                    ("Proformas convertidas:", total_convertidas, "", "Monto total convertido:", f"${monto_convertido:,.2f}"),
                    ("Valor promedio de ventas:", f"${monto_convertido/total_convertidas if total_convertidas > 0 else 0:,.2f}", "", "", ""),
                ]
                
                for i, row_data in enumerate(ventas_data, 4):
                    for j, value in enumerate(row_data, 1):
                        if j in [1, 4]:  # Columnas de etiquetas
                            ws_ventas.cell(row=i, column=j, value=value).font = Font(bold=True)
                        else:
                            ws_ventas.cell(row=i, column=j, value=value)
                
                # Análisis de tiempo de conversión
                ws_ventas.cell(row=8, column=1, value="TIEMPO DE CONVERSIÓN").font = subheader_font
                
                # Calcular estadísticas de tiempo desde emisión hasta conversión
                if ventas_queryset.exists():
                    # Añadir un campo calculado para los días hasta la conversión
                    tiempo_conversion = ventas_queryset.annotate(
                        dias_conversion=ExpressionWrapper(
                            F('updated_at') - F('created_at'),
                            output_field=models.DurationField()
                        )
                    )
                    
                    # Calcular promedios
                    from django.db.models.functions import Cast
                    from django.db.models import DurationField
                    
                    # Extraer datos para análisis
                    tiempos = []
                    for venta in tiempo_conversion:
                        try:
                            # Calcular días entre creación y conversión
                            dias = (venta.updated_at - venta.created_at).days
                            tiempos.append(dias)
                        except Exception as e:
                            logger.warning(f"Error al calcular tiempo de conversión: {str(e)}")
                    
                    # Calcular estadísticas básicas
                    if tiempos:
                        tiempo_promedio = sum(tiempos) / len(tiempos)
                        tiempo_minimo = min(tiempos)
                        tiempo_maximo = max(tiempos)
                        
                        ws_ventas.cell(row=9, column=1, value="Tiempo promedio de conversión:").font = Font(bold=True)
                        ws_ventas.cell(row=9, column=2, value=f"{tiempo_promedio:.1f} días")
                        
                        ws_ventas.cell(row=10, column=1, value="Tiempo mínimo:").font = Font(bold=True)
                        ws_ventas.cell(row=10, column=2, value=f"{tiempo_minimo} días")
                        
                        ws_ventas.cell(row=11, column=1, value="Tiempo máximo:").font = Font(bold=True)
                        ws_ventas.cell(row=11, column=2, value=f"{tiempo_maximo} días")
                else:
                    ws_ventas.cell(row=9, column=1, value="No hay datos de conversión disponibles en el período seleccionado.")
            
            # Análisis por vendedor
            ws_vendedores = wb.create_sheet(title="Por Vendedor")
            
            # Título
            ws_vendedores.merge_cells('A1:E1')
            title_cell = ws_vendedores.cell(row=1, column=1, value="DESEMPEÑO POR VENDEDOR")
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center")
            
            # Encabezados para tabla de vendedores
            headers = ["Vendedor", "Proformas", "Monto Total", "Promedio", "Aprobadas", "Tasa Aprobación", "Tasa Conversión"]
            for col_num, header in enumerate(headers, 1):
                cell = ws_vendedores.cell(row=3, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                ws_vendedores.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 2)
            
            # Obtener estadísticas por vendedor
            from django.contrib.auth import get_user_model
            User = get_user_model()
            vendedor_stats = queryset.values(
                'created_by_id'
            ).annotate(
                cantidad=Count('id'),
                total=Sum('total'),
                promedio=ExpressionWrapper(Sum('total') / Count('id'), output_field=FloatField()),
                aprobadas=Count('id', filter=models.Q(estado='aprobada')),
                convertidas=Count('id', filter=models.Q(estado='convertida')),
            ).order_by('-total')
            
            # Añadir información del vendedor y tasas
            for stats in vendedor_stats:
                if stats['created_by_id']:
                    try:
                        vendedor = User.objects.get(pk=stats['created_by_id'])
                        stats['vendedor_nombre'] = vendedor.get_full_name() or vendedor.username
                    except User.DoesNotExist:
                        stats['vendedor_nombre'] = f"ID: {stats['created_by_id']}"
                else:
                    stats['vendedor_nombre'] = "Sistema"
                
                cantidad = stats['cantidad']
                stats['tasa_aprobacion'] = (stats['aprobadas'] / cantidad * 100) if cantidad > 0 else 0
                stats['tasa_conversion'] = (stats['convertidas'] / cantidad * 100) if cantidad > 0 else 0
            
            # Añadir datos a la tabla
            row = 4
            for vendedor_stat in vendedor_stats:
                ws_vendedores.cell(row=row, column=1, value=vendedor_stat['vendedor_nombre'])
                ws_vendedores.cell(row=row, column=2, value=vendedor_stat['cantidad'])
                
                # Monto total
                cell = ws_vendedores.cell(row=row, column=3, value=vendedor_stat['total'])
                cell.number_format = '#,##0.00'
                
                # Promedio
                cell = ws_vendedores.cell(row=row, column=4, value=vendedor_stat['promedio'])
                cell.number_format = '#,##0.00'
                
                # Cantidad aprobadas
                ws_vendedores.cell(row=row, column=5, value=vendedor_stat['aprobadas'])
                
                # Tasa de aprobación
                cell = ws_vendedores.cell(row=row, column=6, value=vendedor_stat['tasa_aprobacion'] / 100)
                cell.number_format = '0.00%'
                
                # Tasa de conversión
                cell = ws_vendedores.cell(row=row, column=7, value=vendedor_stat['tasa_conversion'] / 100)
                cell.number_format = '0.00%'
                
                row += 1
            
            # Gráfico comparativo de vendedores
            if len(vendedor_stats) > 0:
                chart5 = BarChart()
                chart5.title = "Comparativa de Vendedores"
                chart5.style = 10
                chart5.x_axis.title = "Vendedor"
                chart5.y_axis.title = "Monto Total"
                
                data = Reference(ws_vendedores, min_col=3, min_row=3, max_row=min(row-1, 10), max_col=3)  # Limitar a 10 vendedores
                cats = Reference(ws_vendedores, min_col=1, min_row=4, max_row=min(row-1, 10))
                chart5.add_data(data, titles_from_data=True)
                chart5.set_categories(cats)
                chart5.shape = 4
                chart5.height = 15
                chart5.width = 20
                
                ws_vendedores.add_chart(chart5, "I4")
            
            # Crear respuesta HTTP
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Nombre del archivo con fecha
            fecha_str = timezone.now().strftime("%Y%m%d_%H%M")
            filename = f"reporte_proformas_{fecha_str}.xlsx"
            response["Content-Disposition"] = f"attachment; filename={filename}"
            
            # Guardar y retornar
            logger.info(f"Generando reporte estadístico de proformas. Usuario: {request.user.username}")
            wb.save(response)
            return response
            
        except Exception as e:
            logger.exception(f"Error al generar reporte estadístico: {str(e)}")
            return Response(
                {"error": f"Error al generar reporte: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """Obtener estadísticas para el dashboard de proformas"""
        try:
            # Filtrar por rango de fechas si se proporciona
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            estado_filter = request.query_params.get('estado')
            cliente_id = request.query_params.get('cliente_id')
            min_total = request.query_params.get('min_total')
            max_total = request.query_params.get('max_total')
            
            queryset = self.get_queryset()
            
            # Aplicar filtros si existen
            if start_date:
                queryset = queryset.filter(fecha_emision__gte=start_date)
            if end_date:
                queryset = queryset.filter(fecha_emision__lte=end_date)
            if estado_filter:
                # Permitir filtrar por múltiples estados
                estados = [estado.strip() for estado in estado_filter.split(',')]
                queryset = queryset.filter(estado__in=estados)
            if cliente_id:
                queryset = queryset.filter(cliente_id=cliente_id)
            if min_total:
                queryset = queryset.filter(total__gte=min_total)
            if max_total:
                queryset = queryset.filter(total__lte=max_total)
            
            # Registrar información de filtrado
            logger.info(
                f"Dashboard filtrado: fechas={start_date}~{end_date}, "
                f"estado={estado_filter}, cliente={cliente_id}, "
                f"total={min_total}~{max_total}, resultados={queryset.count()}"
            )
            
            # Estadísticas por estado
            estado_stats = {}
            for estado, label in Proforma.ESTADO_CHOICES:
                count = queryset.filter(estado=estado).count()
                total = queryset.filter(estado=estado).values_list('total', flat=True)
                estado_stats[estado] = {
                    'count': count,
                    'total': sum(total) if total else 0,
                    'label': label
                }
            
            # Estadísticas por cliente (top 5)
            cliente_stats = []
            try:
                top_clientes = queryset.values('cliente__id', 'cliente__nombre').annotate(
                    count=models.Count('id'),
                    total=models.Sum('total')
                ).order_by('-total')[:5]
                
                for cliente in top_clientes:
                    cliente_stats.append({
                        'id': cliente['cliente__id'],
                        'nombre': cliente['cliente__nombre'],
                        'count': cliente['count'],
                        'total': cliente['total'] or 0
                    })
            except Exception as e:
                logger.error(f"Error al procesar estadísticas por cliente: {str(e)}")
                # Proporcionar datos vacíos en caso de error
                cliente_stats = []
            
            # Estadísticas por mes
            from django.db.models.functions import TruncMonth
            
            mes_stats = []
            try:
                por_mes = queryset.annotate(
                    mes=TruncMonth('fecha_emision')
                ).values('mes').annotate(
                    count=models.Count('id'),
                    total=models.Sum('total')
                ).order_by('mes')
                
                for mes in por_mes:
                    if mes['mes']:
                        mes_stats.append({
                            'mes': mes['mes'].strftime('%Y-%m'),
                            'count': mes['count'],
                            'total': mes['total'] or 0
                        })
            except Exception as e:
                logger.error(f"Error al procesar estadísticas por mes: {str(e)}")
                # Proporcionar datos vacíos en caso de error
                mes_stats = []
            
            # Obtener proformas recientes (últimas 5)
            proformas_recientes = []
            try:
                recientes = Proforma.objects.select_related('cliente', 'created_by').order_by('-created_at')[:5]
                
                for proforma in recientes:
                    try:
                        # Obtener nombres para el estado
                        estado_label = dict(Proforma.ESTADO_CHOICES).get(proforma.estado, proforma.estado)
                        
                        # Formatear fechas
                        fecha_emision = proforma.fecha_emision.strftime('%d/%m/%Y') if proforma.fecha_emision else ''
                        fecha_vencimiento = proforma.fecha_vencimiento.strftime('%d/%m/%Y') if proforma.fecha_vencimiento else ''
                        
                        # Iniciales del cliente para el avatar
                        cliente_nombre = proforma.cliente.nombre if proforma.cliente else 'N/A'
                        cliente_avatar = ''.join([word[0] for word in cliente_nombre.split()[:2]]) if cliente_nombre != 'N/A' else 'NA'
                        
                        # Vendedor (usuario que creó la proforma)
                        vendedor = f"{proforma.created_by.first_name} {proforma.created_by.last_name}" if proforma.created_by else 'N/A'
                        if vendedor.strip() == '':
                            vendedor = proforma.created_by.username if proforma.created_by else 'N/A'
                        
                        proformas_recientes.append({
                            'id': proforma.numero,  # Usamos 'numero' como ID para la interfaz
                            'numero': proforma.numero,
                            'cliente': cliente_nombre,
                            'clienteAvatar': cliente_avatar,
                            'fecha': fecha_emision,
                            'expira': fecha_vencimiento,
                            'monto': float(proforma.total),
                            'estado': estado_label,
                            'vendedor': vendedor
                        })
                    except Exception as e:
                        logger.error(f"Error al procesar proforma {proforma.id}: {str(e)}")
                        # Continuar con la siguiente proforma en caso de error
            except Exception as e:
                logger.error(f"Error al obtener proformas recientes: {str(e)}")
                # Proporcionar datos vacíos en caso de error
                proformas_recientes = []
            
            # Devolver estadísticas y proformas recientes
            # Crear datos de respuesta
            try:
                total_count = queryset.count()
                total_aprobadas = queryset.filter(estado='aprobada').count()
                
                try:
                    total_monto = float(sum(queryset.values_list('total', flat=True)))
                except Exception:
                    logger.error("Error al calcular suma de montos, usando 0")
                    total_monto = 0.0
                    
                try:
                    tasa_conversion = round((total_aprobadas / total_count) * 100, 1) if total_count > 0 else 0
                except Exception:
                    logger.error("Error al calcular tasa de conversión, usando 0")
                    tasa_conversion = 0.0
                
                response_data = {
                    'total_proformas': total_count,
                    'total_monto': total_monto,
                    'por_estado': estado_stats,
                    'por_cliente': cliente_stats,
                    'por_mes': mes_stats,
                    'proformasRecientes': proformas_recientes,
                    # Añadir campos totales para las estadísticas principales
                    'totalStats': {
                        'totalProformas': total_count,
                        'proformasAprobadas': total_aprobadas,
                        'tasaConversion': tasa_conversion,
                        'montoTotal': total_monto
                    }
                }
            except Exception as e:
                logger.error(f"Error al crear datos de respuesta: {str(e)}")
                # Proporcionar estructura mínima en caso de error
                response_data = {
                    'total_proformas': 0,
                    'total_monto': 0.0,
                    'por_estado': {},
                    'por_cliente': [],
                    'por_mes': [],
                    'proformasRecientes': [],
                    'totalStats': {
                        'totalProformas': 0,
                        'proformasAprobadas': 0,
                        'tasaConversion': 0.0,
                        'montoTotal': 0.0
                    }
                }
            
            # Log para depuración
            if settings.DEBUG:
                logger.debug(f"Respuesta dashboard API: {len(str(response_data))} bytes")
            
            return Response(response_data)
            
        except Exception as e:
            logger.error(f"Error en dashboard: {str(e)}")
            return Response(
                {"error": "Error al generar el dashboard"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """
        Exportar proformas a formato Excel
        
        Parámetros de filtrado disponibles:
        - estado: Estado de las proformas ('borrador', 'enviada', etc.) o múltiples estados separados por coma
        - cliente_id: ID del cliente
        - fecha_inicio: Fecha de inicio (formato: YYYY-MM-DD)
        - fecha_fin: Fecha fin (formato: YYYY-MM-DD)
        - order_by: Campo para ordenar (default: -fecha_emision)
        - limit: Número máximo de registros a exportar (default: 5000)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Obtener parámetros de filtrado
            estado = request.query_params.get('estado', None)
            cliente_id = request.query_params.get('cliente_id', None)
            fecha_inicio = request.query_params.get('fecha_inicio', None)
            fecha_fin = request.query_params.get('fecha_fin', None)
            order_by = request.query_params.get('order_by', '-fecha_emision')
            limit = int(request.query_params.get('limit', 5000))
            
            # Construir queryset con filtros
            queryset = Proforma.objects.select_related('cliente', 'empresa').all()
            
            # Aplicar filtros si se proporcionan
            if estado:
                if ',' in estado:
                    estados = [e.strip() for e in estado.split(',')]
                    queryset = queryset.filter(estado__in=estados)
                else:
                    queryset = queryset.filter(estado=estado)
            
            if cliente_id:
                queryset = queryset.filter(cliente_id=cliente_id)
                
            if fecha_inicio:
                queryset = queryset.filter(fecha_emision__gte=fecha_inicio)
                
            if fecha_fin:
                queryset = queryset.filter(fecha_emision__lte=fecha_fin)
            
            # Ordenar resultados
            queryset = queryset.order_by(order_by)
            
            # Limitar cantidad de registros para evitar problemas de memoria
            queryset = queryset[:limit]
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Proformas"
            
            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Definir bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Añadir encabezados
            headers = [
                "Número", "Nombre", "Cliente", "RUC/ID", "Fecha Emisión", 
                "Fecha Vencimiento", "Subtotal", "IVA", "Total", "Estado",
                "Empresa Emisora", "Creada por", "Fecha Creación"
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # Ajustar ancho de columnas según el encabezado
                ws.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 2)
            
            # Añadir datos
            for row_num, proforma in enumerate(queryset, 2):
                # Obtener nombre completo del usuario que creó la proforma
                created_by = ""
                if proforma.created_by:
                    first_name = getattr(proforma.created_by, 'first_name', '')
                    last_name = getattr(proforma.created_by, 'last_name', '')
                    if first_name or last_name:
                        created_by = f"{first_name} {last_name}".strip()
                    else:
                        created_by = proforma.created_by.username
                
                # Formatear fechas
                fecha_emision = proforma.fecha_emision.strftime("%d/%m/%Y") if proforma.fecha_emision else ""
                fecha_vencimiento = proforma.fecha_vencimiento.strftime("%d/%m/%Y") if proforma.fecha_vencimiento else ""
                fecha_creacion = proforma.created_at.strftime("%d/%m/%Y %H:%M") if proforma.created_at else ""
                
                # Formatear montos
                subtotal = float(proforma.subtotal) if proforma.subtotal else 0
                impuesto = float(proforma.impuesto) if proforma.impuesto else 0
                total = float(proforma.total) if proforma.total else 0
                
                # Obtener etiqueta legible del estado
                estado_display = dict(Proforma.ESTADO_CHOICES).get(proforma.estado, proforma.estado)
                
                # Datos de la fila
                row = [
                    proforma.numero,
                    proforma.nombre or "",
                    proforma.cliente.nombre if proforma.cliente else "",
                    proforma.cliente.ruc if proforma.cliente else "",
                    fecha_emision,
                    fecha_vencimiento,
                    subtotal,
                    impuesto,
                    total,
                    estado_display,
                    proforma.empresa.nombre if proforma.empresa else "",
                    created_by,
                    fecha_creacion
                ]
                
                # Escribir datos en la hoja
                for col_num, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.border = thin_border
                    
                    # Alinear numeros a la derecha
                    if isinstance(cell_value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                        
                    # Ajustar ancho de columnas según contenido
                    current_width = ws.column_dimensions[get_column_letter(col_num)].width
                    content_length = len(str(cell_value)) + 2
                    if content_length > current_width:
                        ws.column_dimensions[get_column_letter(col_num)].width = min(50, content_length)
                        
            # Formatear columnas numéricas para moneda
            for col in [7, 8, 9]:  # Subtotal, IVA, Total
                for row in range(2, len(queryset) + 2):
                    cell = ws.cell(row=row, column=col)
                    cell.number_format = '#,##0.00'
            
            # Congelar la primera fila
            ws.freeze_panes = "A2"
            
            # Ajustar configuración de impresión
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            
            # Crear respuesta HTTP con el archivo Excel
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = "attachment; filename=proformas.xlsx"
            
            # Registrar la exportación
            total_registros = queryset.count()
            logger.info(f"Exportando {total_registros} proformas a Excel. Usuario: {request.user.username}")
            
            # Guardar el libro de Excel en la respuesta
            wb.save(response)
            return response
        
        except Exception as e:
            logger.exception(f"Error al exportar proformas a Excel: {str(e)}")
            return Response(
                {"error": f"Error al exportar a Excel: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    @action(detail=True, methods=['get'])
    def exportar_csv(self, request, pk=None):
        """Exportar una proforma a formato CSV"""
        proforma = self.get_object()
        
        # Crear buffer en memoria para el CSV
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Encabezado
        writer.writerow(['Proforma', proforma.numero])
        writer.writerow(['Fecha Emisión', proforma.fecha_emision.strftime('%Y-%m-%d')])
        writer.writerow(['Fecha Vencimiento', proforma.fecha_vencimiento.strftime('%Y-%m-%d')])
        writer.writerow(['Cliente', proforma.cliente.nombre])
        writer.writerow(['RUC Cliente', proforma.cliente.ruc])
        writer.writerow(['Empresa Emisora', proforma.empresa.nombre])
        writer.writerow(['Estado', dict(Proforma.ESTADO_CHOICES)[proforma.estado]])
        writer.writerow([])
        
        # Ítems
        writer.writerow(['Código', 'Descripción', 'Unidad', 'Cantidad', 'Precio Unitario', 'Descuento %', 'Total'])
        
        for item in proforma.items.all():
            writer.writerow([
                item.codigo,
                item.descripcion,
                item.unidad,
                item.cantidad,
                item.precio_unitario,
                item.porcentaje_descuento,
                item.total
            ])
        
        writer.writerow([])
        
        # Totales
        writer.writerow(['', '', '', '', '', 'Subtotal', proforma.subtotal])
        writer.writerow(['', '', '', '', '', f'Impuesto ({proforma.porcentaje_impuesto}%)', proforma.impuesto])
        writer.writerow(['', '', '', '', '', 'Total', proforma.total])
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="proforma_{proforma.numero}.csv"'
        
        return response
        
    @action(detail=True, methods=['get'])
    def exportar_excel_detalle(self, request, pk=None):
        """
        Exportar una proforma específica a Excel con todos sus detalles e ítems
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Obtener la proforma con todos sus datos relacionados
            proforma = self.get_object()
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            
            # Primera hoja: Datos generales de la proforma
            ws = wb.active
            ws.title = f"Proforma {proforma.numero}"
            
            # Estilos
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            subheader_font = Font(bold=True)
            subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:F1')
            title_cell = ws.cell(row=1, column=1, value=f"PROFORMA #{proforma.numero}")
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center")
            
            # Sección: Datos Generales
            ws.cell(row=3, column=1, value="INFORMACIÓN GENERAL").font = subheader_font
            
            # Datos básicos
            datos_basicos = [
                ("Número:", proforma.numero),
                ("Nombre:", proforma.nombre or ""),
                ("Estado:", dict(Proforma.ESTADO_CHOICES).get(proforma.estado, proforma.estado)),
                ("Fecha Emisión:", proforma.fecha_emision.strftime("%d/%m/%Y") if proforma.fecha_emision else ""),
                ("Fecha Vencimiento:", proforma.fecha_vencimiento.strftime("%d/%m/%Y") if proforma.fecha_vencimiento else ""),
                ("Creado por:", f"{proforma.created_by.first_name} {proforma.created_by.last_name}".strip() 
                             if proforma.created_by else ""),
                ("Fecha Creación:", proforma.created_at.strftime("%d/%m/%Y %H:%M") if proforma.created_at else ""),
            ]
            
            for i, (label, value) in enumerate(datos_basicos, 4):
                ws.cell(row=i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=i, column=2, value=value)
            
            # Información del cliente
            ws.cell(row=3, column=4, value="INFORMACIÓN DEL CLIENTE").font = subheader_font
            
            datos_cliente = [
                ("Cliente:", proforma.cliente.nombre if proforma.cliente else ""),
                ("RUC/ID:", proforma.cliente.ruc if proforma.cliente else ""),
                ("Dirección:", proforma.cliente.direccion if proforma.cliente else ""),
                ("Teléfono:", proforma.cliente.telefono if proforma.cliente else ""),
                ("Email:", proforma.cliente.email if proforma.cliente else ""),
                ("Atención a:", proforma.atencion_a or ""),
            ]
            
            for i, (label, value) in enumerate(datos_cliente, 4):
                ws.cell(row=i, column=4, value=label).font = Font(bold=True)
                ws.cell(row=i, column=5, value=value)
            
            # Información de la empresa emisora
            row_start = 12
            ws.cell(row=row_start, column=1, value="INFORMACIÓN DE LA EMPRESA").font = subheader_font
            
            datos_empresa = [
                ("Empresa:", proforma.empresa.nombre if proforma.empresa else ""),
                ("RUC:", proforma.empresa.ruc if proforma.empresa else ""),
                ("Dirección:", proforma.empresa.direccion if proforma.empresa else ""),
                ("Teléfono:", proforma.empresa.telefono if proforma.empresa else ""),
            ]
            
            for i, (label, value) in enumerate(datos_empresa, row_start + 1):
                ws.cell(row=i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=i, column=2, value=value)
            
            # Condiciones comerciales
            ws.cell(row=row_start, column=4, value="CONDICIONES COMERCIALES").font = subheader_font
            
            datos_condiciones = [
                ("Condiciones de Pago:", proforma.condiciones_pago or ""),
                ("Tiempo de Entrega:", proforma.tiempo_entrega or ""),
                ("Tipo Contratación:", proforma.tipo_contratacion.nombre if proforma.tipo_contratacion else ""),
                ("Porcentaje Impuesto:", f"{proforma.porcentaje_impuesto}%"),
            ]
            
            for i, (label, value) in enumerate(datos_condiciones, row_start + 1):
                ws.cell(row=i, column=4, value=label).font = Font(bold=True)
                ws.cell(row=i, column=5, value=value)
            
            # Sección: Items
            row_items = row_start + 7
            ws.cell(row=row_items, column=1, value="DETALLE DE PRODUCTOS/SERVICIOS").font = subheader_font
            ws.merge_cells(f'A{row_items}:G{row_items}')
            
            # Encabezados de tabla de ítems
            headers = ["#", "Código", "Descripción", "Unidad", "Cantidad", "Precio Unit.", "Descuento %", "Total"]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_items + 2, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 2)
            
            # Datos de ítems
            items = proforma.items.all().order_by('orden')
            
            for i, item in enumerate(items, 1):
                row = row_items + 2 + i
                valores = [
                    i,
                    item.codigo,
                    item.descripcion,
                    item.unidad,
                    float(item.cantidad),
                    float(item.precio_unitario),
                    float(item.porcentaje_descuento),
                    float(item.total),
                ]
                
                for col_num, valor in enumerate(valores, 1):
                    cell = ws.cell(row=row, column=col_num, value=valor)
                    cell.border = thin_border
                    
                    # Alineación para números
                    if isinstance(valor, (int, float)) and col_num > 3:
                        cell.alignment = Alignment(horizontal="right")
                        
                    # Formato de moneda para campos específicos
                    if col_num in [6, 8]:  # Precio unitario y total
                        cell.number_format = '#,##0.00'
                    
                    # Formato de números para cantidad
                    if col_num == 5:  # Cantidad
                        if valor.is_integer():
                            cell.number_format = '0'
                        else:
                            cell.number_format = '#,##0.00'
                    
                    # Formato para descuento
                    if col_num == 7:  # Descuento
                        cell.number_format = '0.00"%"'
            
            # Sección: Totales
            row_totales = row_items + 4 + len(items)
            
            totales = [
                ("Subtotal:", float(proforma.subtotal)),
                (f"IVA ({proforma.porcentaje_impuesto}%):", float(proforma.impuesto)),
                ("TOTAL:", float(proforma.total)),
            ]
            
            for i, (label, valor) in enumerate(totales):
                row = row_totales + i
                
                # Etiqueta
                label_cell = ws.cell(row=row, column=6, value=label)
                label_cell.font = Font(bold=True)
                label_cell.alignment = Alignment(horizontal="right")
                
                # Valor
                valor_cell = ws.cell(row=row, column=8, value=valor)
                valor_cell.number_format = '#,##0.00'
                valor_cell.font = Font(bold=True if i == 2 else False)  # Negrita para el total
                valor_cell.alignment = Alignment(horizontal="right")
                
                # Dar formato especial al total
                if i == 2:
                    valor_cell.font = Font(bold=True, size=12)
                    ws.row_dimensions[row].height = 20
            
            # Agregar notas si existen
            if proforma.notas:
                row_notas = row_totales + 5
                ws.cell(row=row_notas, column=1, value="NOTAS:").font = Font(bold=True)
                ws.merge_cells(f'A{row_notas}:A{row_notas}')
                
                nota_cell = ws.cell(row=row_notas, column=2, value=proforma.notas)
                nota_cell.alignment = Alignment(wrap_text=True)
                ws.merge_cells(f'B{row_notas}:H{row_notas}')
                ws.row_dimensions[row_notas].height = 60  # Altura ajustable para notas largas
            
            # Segunda hoja: Historial de cambios
            if proforma.historial.exists():
                ws_historial = wb.create_sheet(title="Historial")
                
                # Título
                ws_historial.merge_cells('A1:E1')
                title_cell = ws_historial.cell(row=1, column=1, value=f"HISTORIAL DE CAMBIOS - PROFORMA #{proforma.numero}")
                title_cell.font = title_font
                title_cell.alignment = Alignment(horizontal="center")
                
                # Encabezados
                headers = ["Fecha", "Acción", "Estado Anterior", "Estado Nuevo", "Usuario", "Notas"]
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws_historial.cell(row=3, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    ws_historial.column_dimensions[get_column_letter(col_num)].width = max(15, len(header) + 2)
                
                # Datos del historial
                historial = proforma.historial.all().order_by('-created_at')
                
                for i, registro in enumerate(historial, 4):
                    # Formatear acción
                    accion_display = dict(ProformaHistorial.ACCION_CHOICES).get(registro.accion, registro.accion)
                    
                    # Formatear estados
                    estado_anterior = dict(Proforma.ESTADO_CHOICES).get(registro.estado_anterior, registro.estado_anterior)
                    estado_nuevo = dict(Proforma.ESTADO_CHOICES).get(registro.estado_nuevo, registro.estado_nuevo)
                    
                    # Formatear usuario
                    usuario = ""
                    if registro.created_by:
                        first_name = getattr(registro.created_by, 'first_name', '')
                        last_name = getattr(registro.created_by, 'last_name', '')
                        if first_name or last_name:
                            usuario = f"{first_name} {last_name}".strip()
                        else:
                            usuario = registro.created_by.username
                    
                    row = [
                        registro.created_at.strftime("%d/%m/%Y %H:%M") if registro.created_at else "",
                        accion_display,
                        estado_anterior,
                        estado_nuevo,
                        usuario,
                        registro.notas or ""
                    ]
                    
                    for col_num, valor in enumerate(row, 1):
                        cell = ws_historial.cell(row=i, column=col_num, value=valor)
                        cell.border = thin_border
                        
                        # Ajustar ancho de columnas según contenido
                        if valor:
                            current_width = ws_historial.column_dimensions[get_column_letter(col_num)].width
                            content_length = min(100, len(str(valor)) + 2)
                            if content_length > current_width:
                                ws_historial.column_dimensions[get_column_letter(col_num)].width = content_length
                                
                        # Formato para notas largas
                        if col_num == 6 and valor:  # Notas
                            cell.alignment = Alignment(wrap_text=True)
                            if len(valor) > 50:
                                ws_historial.row_dimensions[i].height = 30
            
            # Crear respuesta HTTP
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"proforma_{proforma.numero}.xlsx"
            response["Content-Disposition"] = f"attachment; filename={filename}"
            
            # Guardar y retornar
            logger.info(f"Exportando proforma {proforma.numero} a Excel detallado. Usuario: {request.user.username}")
            wb.save(response)
            return response
            
        except Exception as e:
            logger.exception(f"Error al exportar proforma {pk} a Excel detallado: {str(e)}")
            return Response(
                {"error": f"Error al exportar a Excel: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def exportar_pdf(self, request, pk=None):
        """Exportar una proforma a formato PDF"""
        from django.template.loader import render_to_string
        from weasyprint import HTML, CSS
        from django.core.files.base import ContentFile
        import tempfile
        from rest_framework_simplejwt.tokens import AccessToken, TokenError
        
        # Manejar token pasado como parámetro GET para compatibilidad con iframe
        token = request.GET.get('token')
        if token and not request.user.is_authenticated:
            try:
                # Decodificar el token y obtener el usuario
                access_token = AccessToken(token)
                user_id = access_token['user_id']
                from django.contrib.auth import get_user_model
                User = get_user_model()
                user = User.objects.get(id=user_id)
                # Simular autenticación asignando el usuario a la solicitud
                request.user = user
            except (TokenError, User.DoesNotExist) as e:
                # Si el token no es válido, continuamos con la autenticación estándar
                pass
        
        # Obtener la proforma (esto ya aplica los permisos)
        proforma = self.get_object()
        items = proforma.items.all().order_by('orden')
        
        # Obtener configuración global de proformas
        configuracion = ConfiguracionProforma.objects.first()
        if not configuracion:
            configuracion = ConfiguracionProforma.objects.create()
        
        # Preparar contexto para la plantilla
        context = {
            'proforma': proforma,
            'items': items,
            'cliente': proforma.cliente,
            'empresa': proforma.empresa,
            'config': configuracion,
            'estado': dict(Proforma.ESTADO_CHOICES)[proforma.estado],
            'BASE_URL': request.build_absolute_uri('/').rstrip('/')
        }
        
        # Generar HTML a partir de la plantilla
        html_string = render_to_string('proformas/pdf_template.html', context)
        
        # Crear archivo PDF temporal
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as output:
            try:
                # Generar PDF con WeasyPrint
                html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
                css = CSS(string='''
                    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700&family=Open+Sans:wght@300;400;600&display=swap');
                    
                    @page {
                        size: letter;
                        margin: 1.5cm;
                        @bottom-right {
                            content: "Página " counter(page) " de " counter(pages);
                            font-size: 10px;
                            color: #6B7280;
                        }
                    }
                    
                    /* No se necesita mucho CSS aquí ya que la mayoría está incrustado en el template */
                    /* Esto solo se usa para sobreescribir propiedades si es necesario */
                    
                    .num-cell {
                        text-align: right !important;
                    }
                    
                    /* Un pequeño ajuste para asegurar márgenes en el documento impreso */
                    body {
                        margin: 0;
                        padding: 0;
                    }
                    
                    /* Asegurar que las tablas tengan el ancho correcto en PDF */
                    table { 
                        table-layout: fixed;
                    }
                    
                    /* Mejora las líneas de firma */
                    .signature-line {
                        margin-top: 3em;
                        border-top: 1px solid #000 !important;
                    }
                ''')
                
                html.write_pdf(target=output.name, stylesheets=[css])
                
                # Devolver el archivo PDF como respuesta HTTP
                with open(output.name, 'rb') as pdf_file:
                    response = HttpResponse(pdf_file.read(), content_type='application/pdf')
                    
                    # Por defecto abrir en navegador en lugar de forzar descarga
                    inline_param = request.GET.get('inline', 'true')
                    if inline_param.lower() == 'true':
                        response['Content-Disposition'] = f'inline; filename="proforma_{proforma.numero}.pdf"'
                    else:
                        response['Content-Disposition'] = f'attachment; filename="proforma_{proforma.numero}.pdf"'
                    
                    return response
            except Exception as e:
                import traceback
                print(f"Error generando PDF: {str(e)}")
                print(traceback.format_exc())
                return Response(
                    {"error": f"Error al generar PDF: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
    
    @action(detail=False, methods=['get'])
    def buscar_productos(self, request):
        """Buscar productos para agregar a una proforma"""
        search_term = request.query_params.get('term', '')
        source = request.query_params.get('source', 'all')
        
        if not search_term or len(search_term) < 2:
            return Response([])
        
        results = []
        
        # Buscar en productos ofertados
        if source in ['all', 'ofertados']:
            ofertados = ProductoOfertado.objects.filter(
                Q(code__icontains=search_term) | 
                Q(nombre__icontains=search_term) |
                Q(descripcion__icontains=search_term)
            ).filter(is_active=True)[:10]
            
            for producto in ofertados:
                results.append(BusquedaProductosSerializer.from_producto_ofertado(producto))
        
        # Buscar en productos disponibles
        if source in ['all', 'disponibles']:
            disponibles = ProductoDisponible.objects.filter(
                Q(code__icontains=search_term) | 
                Q(nombre__icontains=search_term)
            ).filter(is_active=True)[:10]
            
            for producto in disponibles:
                results.append(BusquedaProductosSerializer.from_producto_disponible(producto))
        
        return Response(results)
    
    @action(detail=False, methods=['get'])
    def obtener_configuracion(self, request):
        """Obtener la configuración actual de proformas"""
        config = ConfiguracionProforma.objects.first()
        
        if not config:
            # Crear configuración por defecto si no existe
            config = ConfiguracionProforma.objects.create()
        
        serializer = ConfiguracionProformaSerializer(config)
        return Response(serializer.data)


class ProformaItemViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar ítems de proformas
    """
    queryset = ProformaItem.objects.select_related(
        'proforma', 'producto_ofertado', 'producto_disponible'
    ).all()
    serializer_class = ProformaItemSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = ProformaItemFilter
    ordering_fields = ['orden', 'codigo', 'precio_unitario', 'total']
    ordering = ['proforma', 'orden']
    pagination_class = LargeResultsSetPagination
    
    def perform_destroy(self, instance):
        """Cuando se elimina un ítem, recalcular los totales de la proforma"""
        proforma = instance.proforma
        super().perform_destroy(instance)
        proforma.calcular_montos()
        proforma.save(update_fields=['subtotal', 'impuesto', 'total'])


class ProformaHistorialViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para consultar el historial de proformas (solo lectura)
    """
    queryset = ProformaHistorial.objects.select_related(
        'proforma', 'created_by'
    ).all()
    serializer_class = ProformaHistorialSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['proforma', 'accion', 'estado_anterior', 'estado_nuevo', 'created_by']
    ordering_fields = ['created_at', 'accion']
    ordering = ['-created_at']
    pagination_class = StandardResultsSetPagination


class ConfiguracionProformaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar la configuración de proformas
    """
    queryset = ConfiguracionProforma.objects.all()
    serializer_class = ConfiguracionProformaSerializer
    permission_classes = [IsAuthenticated]
    
    def get_object(self):
        """Asegurar que siempre se devuelva la configuración actual"""
        config = ConfiguracionProforma.objects.first()
        
        if not config:
            # Crear configuración por defecto si no existe
            config = ConfiguracionProforma.objects.create()
        
        self.check_object_permissions(self.request, config)
        return config
    
    def list(self, request, *args, **kwargs):
        """Devolver siempre la única configuración existente"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
    
    def create(self, request, *args, **kwargs):
        """Sobreescribir create para actualizar la configuración existente"""
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


@api_view(['GET'])
def stats_dashboard(request):
    """
    Endpoint para obtener estadísticas agregadas para el dashboard
    
    Parámetros de filtrado disponibles:
    - fecha_inicio: Fecha de inicio (formato: YYYY-MM-DD)
    - fecha_fin: Fecha fin (formato: YYYY-MM-DD)
    - periodo: Tipo de agrupación temporal ('day', 'week', 'month', 'year')
    - cliente_id: ID del cliente para filtrar
    """
    # Obtener parámetros de filtrado
    fecha_inicio = request.query_params.get('fecha_inicio')
    fecha_fin = request.query_params.get('fecha_fin')
    periodo = request.query_params.get('periodo', 'month')
    cliente_id = request.query_params.get('cliente_id')
    
    # Base queryset
    queryset = Proforma.objects.all()
    
    # Aplicar filtros de fecha
    if fecha_inicio:
        queryset = queryset.filter(fecha_emision__gte=fecha_inicio)
    else:
        # Por defecto, último año
        fecha_inicio = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        queryset = queryset.filter(fecha_emision__gte=fecha_inicio)
        
    if fecha_fin:
        queryset = queryset.filter(fecha_emision__lte=fecha_fin)
        
    # Filtrar por cliente
    if cliente_id:
        queryset = queryset.filter(cliente_id=cliente_id)
    
    # Estadísticas por estado
    por_estado = list(queryset.values('estado') \
                     .annotate(
                        label=F('estado'),
                        cantidad=Count('id'), 
                        monto=Sum('total')
                     ).order_by('estado'))
    
    # Transformar etiquetas de estado para legibilidad
    estado_labels = dict(Proforma.ESTADO_CHOICES)
    for item in por_estado:
        item['label'] = estado_labels.get(item['estado'], item['estado'])
        
    # Estadísticas de conversión
    total_proformas = queryset.count()
    enviadas = queryset.filter(estado='enviada').count()
    aprobadas = queryset.filter(estado='aprobada').count()
    rechazadas = queryset.filter(estado='rechazada').count()
    convertidas = queryset.filter(estado='convertida').count()
    
    tasa_envio = (enviadas / total_proformas * 100) if total_proformas > 0 else 0
    tasa_aprobacion = (aprobadas / enviadas * 100) if enviadas > 0 else 0
    tasa_conversion = (convertidas / aprobadas * 100) if aprobadas > 0 else 0
    tasa_rechazo = (rechazadas / enviadas * 100) if enviadas > 0 else 0
    
    # Estadísticas por tiempo
    if periodo == 'day':
        trunc_func = TruncDay('fecha_emision')
        format_str = '%Y-%m-%d'
    elif periodo == 'week':
        trunc_func = TruncWeek('fecha_emision')
        format_str = '%Y-%U'  # Año-Semana
    elif periodo == 'year':
        # Para años, usamos Extract en lugar de Trunc
        por_tiempo = list(queryset.annotate(
                periodo=Extract('fecha_emision', 'year')
            ).values('periodo')
            .annotate(
                cantidad=Count('id'),
                monto=Sum('total'),
                aprobadas=Count('id', filter=Q(estado='aprobada')),
                convertidas=Count('id', filter=Q(estado='convertida'))
            ).order_by('periodo'))
            
        for item in por_tiempo:
            # Añadir tasa de conversión
            item['tasa_conversion'] = (item['convertidas'] / item['aprobadas'] * 100) if item['aprobadas'] > 0 else 0
            item['fecha'] = str(int(item['periodo']))  # Convertir a string para serializar
    else:
        # Por defecto, agrupar por mes
        trunc_func = TruncMonth('fecha_emision')
        format_str = '%Y-%m'
        
    # Si no es agrupación por año, usar TruncX
    if periodo != 'year':
        por_tiempo = list(queryset.annotate(
                periodo=trunc_func
            ).values('periodo')
            .annotate(
                cantidad=Count('id'),
                monto=Sum('total'),
                aprobadas=Count('id', filter=Q(estado='aprobada')),
                convertidas=Count('id', filter=Q(estado='convertida'))
            ).order_by('periodo'))
            
        for item in por_tiempo:
            # Añadir tasa de conversión y formatear fecha
            item['tasa_conversion'] = (item['convertidas'] / item['aprobadas'] * 100) if item['aprobadas'] > 0 else 0
            item['fecha'] = item['periodo'].strftime(format_str)
    
    # Top clientes por monto
    top_clientes = list(queryset.values(
            'cliente__id', 'cliente__nombre', 'cliente__ruc'
        ).annotate(
            cantidad=Count('id'),
            monto=Sum('total')
        ).order_by('-monto')[:10])
    
    # Resumen total
    monto_total = sum(item['monto'] for item in por_estado)
    monto_promedio = monto_total / total_proformas if total_proformas > 0 else 0
    
    # Crear respuesta
    response_data = {
        'resumen': {
            'total_proformas': total_proformas,
            'monto_total': monto_total,
            'monto_promedio': monto_promedio,
            'tasa_envio': tasa_envio,
            'tasa_aprobacion': tasa_aprobacion, 
            'tasa_conversion': tasa_conversion,
            'tasa_rechazo': tasa_rechazo
        },
        'por_estado': por_estado,
        'por_tiempo': por_tiempo,
        'top_clientes': top_clientes
    }
    
    return Response(response_data)