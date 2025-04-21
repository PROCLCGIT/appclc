"""
Comando para exportar datos de proformas en diferentes formatos.

Este comando permite exportar datos de proformas a formatos como JSON, CSV o Excel,
útil para respaldos, migración de datos o análisis externos.
"""
import os
import json
import csv
import logging
from datetime import datetime
from django.core.management.base import BaseCommand
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Prefetch, Count, Sum, F, Value, CharField
from django.db.models.functions import Concat
from django.conf import settings

from proformas.models import Proforma, ItemProforma

logger = logging.getLogger(__name__)

class Command(BaseCommand):
    help = 'Exporta datos de proformas en diferentes formatos (JSON, CSV, Excel)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--format',
            type=str,
            choices=['json', 'csv', 'excel'],
            default='json',
            help='Formato de exportación (json, csv, excel)'
        )
        parser.add_argument(
            '--output',
            type=str,
            help='Ruta de archivo de salida (opcional, por defecto se genera automáticamente)'
        )
        parser.add_argument(
            '--items',
            action='store_true',
            help='Incluir items de proforma en la exportación'
        )
        parser.add_argument(
            '--filter',
            type=str,
            help='Filtro por estado de proforma (pendiente, aprobada, etc.)'
        )
        parser.add_argument(
            '--start-date',
            type=str,
            help='Fecha inicial para filtrar (formato YYYY-MM-DD)'
        )
        parser.add_argument(
            '--end-date',
            type=str,
            help='Fecha final para filtrar (formato YYYY-MM-DD)'
        )
        parser.add_argument(
            '--limit',
            type=int,
            help='Número máximo de proformas a exportar'
        )

    def handle(self, *args, **options):
        format_type = options['format']
        output_path = options['output']
        include_items = options['items']
        filter_estado = options['filter']
        start_date = options['start_date']
        end_date = options['end_date']
        limit = options['limit']
        
        self.stdout.write(self.style.NOTICE('Iniciando exportación de proformas...'))
        
        # Construir queryset con filtros aplicados
        queryset = self.get_filtered_queryset(filter_estado, start_date, end_date, include_items, limit)
        
        # Generar ruta de archivo si no se especificó
        if not output_path:
            output_path = self.generate_output_path(format_type)
        
        # Crear directorio si no existe
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Exportar según el formato
        if format_type == 'json':
            self.export_to_json(queryset, output_path, include_items)
        elif format_type == 'csv':
            self.export_to_csv(queryset, output_path, include_items)
        elif format_type == 'excel':
            self.export_to_excel(queryset, output_path, include_items)
        
        self.stdout.write(self.style.SUCCESS(f'Exportación completada: {output_path}'))

    def get_filtered_queryset(self, filter_estado, start_date, end_date, include_items, limit):
        """Obtiene queryset de proformas con filtros aplicados"""
        # Consulta base optimizada
        queryset = Proforma.objects.all().select_related('cliente')
        
        # Prefetch de items si es necesario
        if include_items:
            queryset = queryset.prefetch_related(
                Prefetch('items', queryset=ItemProforma.objects.all())
            )
        
        # Aplicar filtros
        if filter_estado:
            queryset = queryset.filter(estado=filter_estado)
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_emision__gte=start_date)
            except ValueError:
                self.stdout.write(self.style.WARNING(f'Formato de fecha inicial incorrecto: {start_date}. Se omitirá este filtro.'))
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(fecha_emision__lte=end_date)
            except ValueError:
                self.stdout.write(self.style.WARNING(f'Formato de fecha final incorrecto: {end_date}. Se omitirá este filtro.'))
        
        # Aplicar límite si se especificó
        if limit and limit > 0:
            queryset = queryset[:limit]
            
        return queryset

    def generate_output_path(self, format_type):
        """Genera una ruta de archivo de salida basada en la fecha actual"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_dir = os.path.join(settings.BASE_DIR, 'exports', 'proformas')
        os.makedirs(export_dir, exist_ok=True)
        return os.path.join(export_dir, f'proformas_export_{timestamp}.{format_type}')

    def export_to_json(self, queryset, output_path, include_items):
        """Exporta proformas a formato JSON"""
        self.stdout.write('Exportando a JSON...')
        
        # Contadores para estadísticas
        total_proformas = 0
        total_items = 0
        
        # Preparar estructura de datos
        export_data = {
            'metadata': {
                'date': datetime.now().isoformat(),
                'format': 'json',
                'version': '1.0'
            },
            'proformas': []
        }
        
        # Procesar cada proforma
        for proforma in queryset:
            proforma_data = {
                'id': proforma.id,
                'numero': proforma.numero,
                'nombre': proforma.nombre,
                'cliente': {
                    'id': proforma.cliente.id if proforma.cliente else None,
                    'nombre': proforma.cliente.nombre if proforma.cliente else None,
                    'ruc': proforma.cliente.ruc if proforma.cliente else None
                } if proforma.cliente else None,
                'fecha_emision': proforma.fecha_emision.isoformat() if proforma.fecha_emision else None,
                'fecha_validez': proforma.fecha_validez.isoformat() if proforma.fecha_validez else None,
                'estado': proforma.estado,
                'subtotal': float(proforma.subtotal) if proforma.subtotal is not None else None,
                'iva': float(proforma.iva) if proforma.iva is not None else None,
                'total': float(proforma.total) if proforma.total is not None else None,
                'notas': proforma.notas
            }
            
            # Incluir items si se solicita
            if include_items and hasattr(proforma, 'items'):
                proforma_data['items'] = []
                for item in proforma.items.all():
                    item_data = {
                        'id': item.id,
                        'codigo': item.codigo,
                        'descripcion': item.descripcion,
                        'unidad': item.unidad,
                        'cantidad': float(item.cantidad) if item.cantidad is not None else None,
                        'precio_unitario': float(item.precio_unitario) if item.precio_unitario is not None else None,
                        'descuento': float(item.descuento) if item.descuento is not None else None,
                        'precio_total': float(item.precio_total) if item.precio_total is not None else None
                    }
                    proforma_data['items'].append(item_data)
                    total_items += 1
            
            export_data['proformas'].append(proforma_data)
            total_proformas += 1
        
        # Agregar estadísticas
        export_data['metadata']['records'] = {
            'proformas': total_proformas,
            'items': total_items if include_items else 'not_included'
        }
        
        # Guardar archivo JSON
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, cls=DjangoJSONEncoder, indent=2, ensure_ascii=False)
        
        self.stdout.write(f'Exportadas {total_proformas} proformas con {total_items} items.')

    def export_to_csv(self, queryset, output_path, include_items):
        """Exporta proformas a formato CSV"""
        self.stdout.write('Exportando a CSV...')
        
        # Separar en dos archivos si se incluyen items
        proforma_path = output_path
        items_path = None
        
        if include_items:
            base_name, ext = os.path.splitext(output_path)
            items_path = f"{base_name}_items{ext}"
        
        # Exportar proformas
        with open(proforma_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Escribir encabezados
            headers = [
                'ID', 'Número', 'Nombre', 'Cliente ID', 'Cliente Nombre', 'Cliente RUC',
                'Fecha Emisión', 'Fecha Validez', 'Estado', 'Subtotal', 'IVA', 'Total', 'Notas'
            ]
            writer.writerow(headers)
            
            # Escribir datos
            total_proformas = 0
            for proforma in queryset:
                row = [
                    proforma.id,
                    proforma.numero,
                    proforma.nombre,
                    proforma.cliente.id if proforma.cliente else None,
                    proforma.cliente.nombre if proforma.cliente else None,
                    proforma.cliente.ruc if proforma.cliente else None,
                    proforma.fecha_emision.strftime('%Y-%m-%d') if proforma.fecha_emision else None,
                    proforma.fecha_validez.strftime('%Y-%m-%d') if proforma.fecha_validez else None,
                    proforma.estado,
                    proforma.subtotal,
                    proforma.iva,
                    proforma.total,
                    proforma.notas
                ]
                writer.writerow(row)
                total_proformas += 1
        
        # Exportar items si se solicita
        if include_items:
            with open(items_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Escribir encabezados
                headers = [
                    'Proforma ID', 'Proforma Número', 'Item ID', 'Código', 'Descripción',
                    'Unidad', 'Cantidad', 'Precio Unitario', 'Descuento', 'Precio Total'
                ]
                writer.writerow(headers)
                
                # Escribir datos
                total_items = 0
                for proforma in queryset:
                    for item in proforma.items.all():
                        row = [
                            proforma.id,
                            proforma.numero,
                            item.id,
                            item.codigo,
                            item.descripcion,
                            item.unidad,
                            item.cantidad,
                            item.precio_unitario,
                            item.descuento,
                            item.precio_total
                        ]
                        writer.writerow(row)
                        total_items += 1
            
            self.stdout.write(f'Exportadas {total_proformas} proformas con {total_items} items.')
            self.stdout.write(f'Archivo de items: {items_path}')
        else:
            self.stdout.write(f'Exportadas {total_proformas} proformas.')

    def export_to_excel(self, queryset, output_path, include_items):
        """Exporta proformas a formato Excel"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            self.stdout.write(self.style.ERROR(
                'Se requieren las bibliotecas pandas y openpyxl para exportar a Excel. '
                'Instale con: pip install pandas openpyxl'
            ))
            return
        
        self.stdout.write('Exportando a Excel...')
        
        # Crear datos para proformas
        proformas_data = []
        items_data = []
        
        for proforma in queryset:
            proforma_row = {
                'ID': proforma.id,
                'Número': proforma.numero,
                'Nombre': proforma.nombre,
                'Cliente ID': proforma.cliente.id if proforma.cliente else None,
                'Cliente': proforma.cliente.nombre if proforma.cliente else None,
                'RUC': proforma.cliente.ruc if proforma.cliente else None,
                'Fecha Emisión': proforma.fecha_emision,
                'Fecha Validez': proforma.fecha_validez,
                'Estado': proforma.estado,
                'Subtotal': proforma.subtotal,
                'IVA': proforma.iva,
                'Total': proforma.total,
                'Notas': proforma.notas
            }
            proformas_data.append(proforma_row)
            
            # Recopilar items si se solicita
            if include_items:
                for item in proforma.items.all():
                    item_row = {
                        'Proforma ID': proforma.id,
                        'Proforma Número': proforma.numero,
                        'Item ID': item.id,
                        'Código': item.codigo,
                        'Descripción': item.descripcion,
                        'Unidad': item.unidad,
                        'Cantidad': item.cantidad,
                        'Precio Unitario': item.precio_unitario,
                        'Descuento (%)': item.descuento,
                        'Precio Total': item.precio_total
                    }
                    items_data.append(item_row)
        
        # Crear DataFrame
        df_proformas = pd.DataFrame(proformas_data)
        
        # Crear libro y hojas
        wb = Workbook()
        ws_proformas = wb.active
        ws_proformas.title = "Proformas"
        
        # Agregar datos de proformas
        for r in dataframe_to_rows(df_proformas, index=False, header=True):
            ws_proformas.append(r)
        
        # Dar formato a la hoja de proformas
        self.format_excel_sheet(ws_proformas)
        
        # Crear hoja de items si se solicita
        if include_items and items_data:
            df_items = pd.DataFrame(items_data)
            ws_items = wb.create_sheet(title="Items")
            
            for r in dataframe_to_rows(df_items, index=False, header=True):
                ws_items.append(r)
            
            # Dar formato a la hoja de items
            self.format_excel_sheet(ws_items)
        
        # Guardar archivo
        wb.save(output_path)
        
        self.stdout.write(f'Exportadas {len(proformas_data)} proformas' + 
                         (f' con {len(items_data)} items.' if include_items else '.'))

    def format_excel_sheet(self, worksheet):
        """Da formato a una hoja de Excel"""
        # Dar formato al encabezado
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill
        
        # Ajustar anchos de columna
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = min(adjusted_width, 50)  # Límite máximo
        
        # Congelar encabezados
        worksheet.freeze_panes = "A2"