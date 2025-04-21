"""
Utilidades para exportación de datos en diferentes formatos.
Soporta Excel, CSV y PDF para todos los módulos.
"""
import csv
import json
import io
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.template.loader import get_template
from django.utils.timezone import now
from django.utils.text import slugify

class ExportError(Exception):
    """Excepción para errores durante la exportación"""
    pass

def export_to_excel(data, filename=None, sheet_name='Datos', headers=None, header_style=None):
    """
    Exporta datos a Excel con formato mejorado.
    
    Args:
        data: Lista de diccionarios con los datos a exportar
        filename: Nombre del archivo (sin extensión)
        sheet_name: Nombre de la hoja de cálculo
        headers: Diccionario {campo: etiqueta} para personalizar encabezados
        header_style: Estilos adicionales para encabezados
        
    Returns:
        HttpResponse: Respuesta para descargar el archivo Excel
    """
    if not data:
        raise ExportError("No hay datos para exportar")
    
    # Crear nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Preparar encabezados
    fields = list(data[0].keys()) if data else []
    
    if headers:
        # Usar encabezados personalizados
        column_headers = [headers.get(field, field) for field in fields]
    else:
        # Usar nombres de campos como encabezados
        column_headers = [field.replace('_', ' ').title() for field in fields]
    
    # Escribir encabezados
    for col_idx, header in enumerate(column_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        
        # Aplicar estilo base a los encabezados
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(style='thin', color="000000")
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        
        # Aplicar estilos adicionales si se especifican
        if header_style:
            if 'font' in header_style:
                cell.font = header_style['font']
            if 'fill' in header_style:
                cell.fill = header_style['fill']
    
    # Escribir datos
    for row_idx, item in enumerate(data, 2):
        for col_idx, field in enumerate(fields, 1):
            value = item.get(field)
            
            # Formatear valores especiales
            if isinstance(value, datetime.datetime):
                value = value.strftime('%d/%m/%Y %H:%M')
            elif isinstance(value, datetime.date):
                value = value.strftime('%d/%m/%Y')
            elif isinstance(value, bool):
                value = 'Sí' if value else 'No'
            
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Ajustar ancho de columnas
    for col_idx, _ in enumerate(fields, 1):
        column_letter = get_column_letter(col_idx)
        # Calcular ancho basado en el contenido
        max_length = 0
        for row_idx, item in enumerate(data, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Ajustar ancho de columna con un mínimo de 10 y un máximo de 50
        adjusted_width = max(10, min(max_length + 2, 50))
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP
    if filename:
        safe_filename = slugify(filename)
    else:
        safe_filename = f"exportacion_{sheet_name.lower()}_{now().strftime('%Y%m%d')}"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}.xlsx"'
    
    # Guardar el libro en la respuesta
    buffer = io.BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    
    return response

def export_to_csv(data, filename=None, headers=None, delimiter=',', quotechar='"'):
    """
    Exporta datos a CSV.
    
    Args:
        data: Lista de diccionarios con los datos a exportar
        filename: Nombre del archivo (sin extensión)
        headers: Diccionario {campo: etiqueta} para personalizar encabezados
        delimiter: Delimitador de campos
        quotechar: Carácter para encerrar strings
        
    Returns:
        HttpResponse: Respuesta para descargar el archivo CSV
    """
    if not data:
        raise ExportError("No hay datos para exportar")
    
    # Preparar buffer y escritor CSV
    buffer = io.StringIO()
    writer = csv.writer(
        buffer, 
        delimiter=delimiter, 
        quotechar=quotechar, 
        quoting=csv.QUOTE_MINIMAL
    )
    
    # Obtener campos desde el primer elemento
    fields = list(data[0].keys()) if data else []
    
    # Preparar encabezados
    if headers:
        # Usar encabezados personalizados
        column_headers = [headers.get(field, field) for field in fields]
    else:
        # Usar nombres de campos como encabezados
        column_headers = [field.replace('_', ' ').title() for field in fields]
    
    # Escribir encabezados
    writer.writerow(column_headers)
    
    # Escribir datos
    for item in data:
        row = []
        for field in fields:
            value = item.get(field)
            
            # Formatear valores especiales
            if isinstance(value, datetime.datetime):
                value = value.strftime('%d/%m/%Y %H:%M')
            elif isinstance(value, datetime.date):
                value = value.strftime('%d/%m/%Y')
            elif isinstance(value, bool):
                value = 'Sí' if value else 'No'
            elif value is None:
                value = ''
            
            row.append(value)
        
        writer.writerow(row)
    
    # Crear respuesta HTTP
    if filename:
        safe_filename = slugify(filename)
    else:
        safe_filename = f"exportacion_{now().strftime('%Y%m%d')}"
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}.csv"'
    
    # Escribir el contenido del buffer en la respuesta
    buffer.seek(0)
    response.write(buffer.getvalue().encode('utf-8-sig'))  # Incluir BOM para compatibilidad con Excel
    
    return response

def export_to_pdf(data, template_name, filename=None, context=None):
    """
    Exporta datos a PDF usando una plantilla HTML y WeasyPrint.
    
    Args:
        data: Datos a exportar (lista, diccionario, etc.)
        template_name: Ruta a la plantilla HTML
        filename: Nombre del archivo (sin extensión)
        context: Contexto adicional para la plantilla
        
    Returns:
        HttpResponse: Respuesta para descargar el archivo PDF
    """
    try:
        from weasyprint import HTML
        from django.template.loader import render_to_string
    except ImportError:
        raise ImportError("WeasyPrint es requerido para exportar a PDF. Instale con: pip install weasyprint")
    
    # Preparar contexto
    ctx = context or {}
    ctx.update({
        'data': data,
        'date': now().strftime('%d/%m/%Y'),
        'time': now().strftime('%H:%M:%S')
    })
    
    # Renderizar HTML usando la plantilla
    html_string = render_to_string(template_name, ctx)
    
    # Convertir HTML a PDF
    html = HTML(string=html_string)
    pdf_file = html.write_pdf()
    
    # Crear respuesta HTTP
    if filename:
        safe_filename = slugify(filename)
    else:
        safe_filename = f"exportacion_{now().strftime('%Y%m%d')}"
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}.pdf"'
    
    # Escribir el contenido del PDF en la respuesta
    response.write(pdf_file)
    
    return response

def export_to_json(data, filename=None, indent=2):
    """
    Exporta datos a JSON.
    
    Args:
        data: Datos a exportar (lista, diccionario, etc.)
        filename: Nombre del archivo (sin extensión)
        indent: Nivel de indentación para el JSON
        
    Returns:
        HttpResponse: Respuesta para descargar el archivo JSON
    """
    class CustomJSONEncoder(json.JSONEncoder):
        """Encoder personalizado para manejar tipos especiales"""
        def default(self, obj):
            if isinstance(obj, (datetime.datetime, datetime.date)):
                return obj.isoformat()
            return super().default(obj)
    
    # Serializar datos a JSON
    json_data = json.dumps(
        data, 
        cls=CustomJSONEncoder, 
        indent=indent,
        ensure_ascii=False
    )
    
    # Crear respuesta HTTP
    if filename:
        safe_filename = slugify(filename)
    else:
        safe_filename = f"exportacion_{now().strftime('%Y%m%d')}"
    
    response = HttpResponse(content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}.json"'
    
    # Escribir el contenido JSON en la respuesta
    response.write(json_data)
    
    return response

# Utilidad para usar en los endpoints
def create_exporter(view_name, model_verbose_name=None):
    """
    Crea una función para manejar exportaciones en diferentes formatos.
    Para usar en métodos @action de viewsets.
    
    Args:
        view_name: Nombre de la vista/recurso (usado para el nombre del archivo)
        model_verbose_name: Nombre legible del modelo (opcional)
        
    Returns:
        function: Función para manejar la exportación
    """
    def export(self, request, format=None):
        # Determinar formato solicitado
        format = format or request.query_params.get('format', 'xlsx').lower()
        
        # Obtener nombre para el archivo
        resource_name = model_verbose_name or view_name
        filename = request.query_params.get(
            'filename', 
            f"{resource_name}_{now().strftime('%Y%m%d')}"
        )
        
        # Configurar filtros
        queryset = self.filter_queryset(self.get_queryset())
        
        # Serializar datos (sin paginación)
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data
        
        # Exportar según formato
        if format == 'xlsx':
            return export_to_excel(data, filename=filename, sheet_name=resource_name)
        elif format == 'csv':
            return export_to_csv(data, filename=filename)
        elif format == 'json':
            return export_to_json(data, filename=filename)
        elif format == 'pdf':
            # Para PDF necesitamos una plantilla
            # Se debe implementar por módulo según necesidades específicas
            return Response({"error": "Formato PDF no soportado para esta vista"}, status=400)
        else:
            return Response({"error": f"Formato no soportado: {format}"}, status=400)
    
    return export
